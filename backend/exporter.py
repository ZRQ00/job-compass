from sqlalchemy.orm import Session
from database import Job
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import io
from datetime import datetime


def export_to_csv(db: Session) -> str:
    jobs = db.query(Job).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Title", "Company", "Location", "Remote", "Salary Min",
        "Salary Max", "Status", "Priority", "Fit Score", "Source",
        "Applied Date", "Notes", "URL", "Date Posted"
    ])
    for j in jobs:
        writer.writerow([
            j.id, j.title, j.company, j.location, j.is_remote,
            j.salary_min, j.salary_max, j.status, j.priority,
            j.fit_score, j.source,
            j.applied_date.strftime("%Y-%m-%d") if j.applied_date else "",
            j.notes, j.job_url, j.date_posted
        ])
    return output.getvalue()


def export_to_xlsx(db: Session) -> bytes:
    jobs = db.query(Job).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    HEADER_BG = "1a1a2e"
    thin = Side(style="thin", color="dee2e6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_colors = {
        "found": "e3f2fd",
        "applied": "fff3cd",
        "interview": "d4edda",
        "offer": "c3e6cb",
        "rejected": "f8d7da",
        "skipped": "e9ecef",
    }

    headers = [
        "Company", "Title", "Location", "Remote", "Salary Min", "Salary Max",
        "Status", "Priority", "Fit Score", "Source", "Applied Date", "Notes", "URL"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, j in enumerate(jobs, 2):
        bg = status_colors.get(j.status, "FFFFFF")
        values = [
            j.company, j.title, j.location, "Yes" if j.is_remote else "No",
            j.salary_min, j.salary_max, j.status, j.priority, j.fit_score,
            j.source,
            j.applied_date.strftime("%Y-%m-%d") if j.applied_date else "",
            j.notes, j.job_url
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row_idx].height = 20

    col_widths = [20, 30, 20, 8, 12, 12, 12, 15, 10, 12, 14, 25, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
